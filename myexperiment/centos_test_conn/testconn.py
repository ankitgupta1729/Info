import openpyxl
import pyodbc
import pandas as pd
print("Testing Connection")
print("pyodbc drivers:\n",pyodbc.drivers()) 
try:
    book = openpyxl.load_workbook('Details.xlsx')
except FileNotFoundError:
    print('Server details sheet not found!')
sheet = book.active
server=sheet.cell(5,2).value
database=sheet.cell(6,2).value
username=sheet.cell(7,2).value
password=sheet.cell(8,2).value

connectionString = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
conn = pyodbc.connect(connectionString)
SQL_QUERY = "SELECT * from RM_ST_FFA_Control_table1"
data = pd.read_sql(SQL_QUERY,conn)
print(data)
cursor = conn.cursor()
cursor.execute(SQL_QUERY)
row = cursor.fetchone()
print(row)
row = cursor.fetchone()
print(row)
row = cursor.fetchone()
print(row)
row = cursor.fetchone()
print(row)
print("connection established using pyodbc")
